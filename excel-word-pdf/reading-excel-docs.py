import openpyxl
import os

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook['Sheet1']
print(type(sheet))

print(workbook.sheetnames)

print(sheet['A1'])
print(sheet['A1'].value)
print(type(sheet['A1'].value))
print(str(sheet['A1'].value))

print(sheet['B1'].value)
print(sheet['C1'].value)

print(sheet.cell(row=1, column=2))

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=1))
