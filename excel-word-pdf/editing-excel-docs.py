'''
You can view and modify a sheet's name with its "title" member variable.
Changing a cell's value is done using the square brackets, just like changing a value in a list or dictionary.
Changes you make to the workbook object can be saved with the save() method.
'''

import openpyxl

wb = openpyxl.Workbook()
print(wb)

print(wb.sheetnames)

sheet = wb['Sheet']
print(sheet)
print(sheet['A1'].value)

sheet['A1'] = 42
sheet['A2'] = 'Hello'

sheet2 = wb.create_sheet()
print(wb.sheetnames)
print(sheet2.title)
sheet2.title = 'My New Sheet Name'
print(wb.sheetnames)

wb.create_sheet(index=0, title='My Other Sheet')
wb.save('edit-example.xlsx')
